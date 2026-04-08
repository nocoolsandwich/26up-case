from __future__ import annotations

import argparse
from contextlib import contextmanager
from concurrent.futures import ThreadPoolExecutor
import logging
import time
from collections import OrderedDict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Callable

import pandas as pd
from openpyxl import load_workbook
import psycopg
import requests
import tushare as ts
import numpy as np

try:
    import akshare as ak
except ModuleNotFoundError:  # pragma: no cover - 生产环境默认已安装
    ak = None

try:
    from scripts.project_config import PROJECT_ROOT, load_project_config
except ModuleNotFoundError:  # 兼容 python scripts/event_quant_sync.py 直接运行
    from project_config import PROJECT_ROOT, load_project_config

logger = logging.getLogger(__name__)

THEME_CONCEPT_PREFIXES = ("884", "885", "886")

DEFAULT_REGULAR_REQUESTS_PER_MIN = 100
DEFAULT_FEATURE_REQUESTS_PER_MIN = 60
DEFAULT_MAX_WORKERS = 4
DEFAULT_CASE_SHEET_NAME = "案例库"
DEFAULT_CASE_STOCK_BUNDLE_JOB_NAME = "sync_case_stock_bundle"
DEFAULT_CASE_STOCK_BUNDLE_TARGET_KEY = "case_stocks"
DEFAULT_ALL_STOCKS_JOB_NAME = "sync_all_stocks_by_trade_date"
DEFAULT_ALL_STOCKS_TARGET_KEY = "all_stocks"
DEFAULT_ALL_CONCEPTS_JOB_NAME = "sync_all_concepts"
DEFAULT_ALL_CONCEPTS_TARGET_KEY = "all_concepts"
DEFAULT_ALL_CONCEPT_MEMBERS_JOB_NAME = "sync_all_concept_members"
DEFAULT_ALL_CONCEPT_MEMBERS_TARGET_KEY = "all_concept_members"
DEFAULT_TUSHARE_HTTP_URL = "http://lianghua.nanyangqiankun.top"
DEFAULT_DB_DSN = None


def sync_targets() -> OrderedDict[str, dict[str, object]]:
    return OrderedDict(
        [
            ("raw_stock_daily_qfq", {"api": "pro_bar", "pk": ("ts_code", "trade_date")}),
            ("raw_index_daily", {"api": "index_daily", "pk": ("ts_code", "trade_date")}),
            ("raw_ths_concept_daily", {"api": "ths_daily", "pk": ("ts_code", "trade_date")}),
            ("raw_ths_member", {"api": "ths_member", "pk": ("ts_code", "con_code", "mapping_asof_date")}),
            ("raw_daily_basic", {"api": "daily_basic", "pk": ("ts_code", "trade_date")}),
            ("raw_moneyflow", {"api": "moneyflow", "pk": ("ts_code", "trade_date")}),
            ("raw_limit_list_d", {"api": "limit_list_d", "pk": ("ts_code", "trade_date")}),
        ]
    )


def table_primary_keys() -> dict[str, tuple[str, ...]]:
    targets = {name: tuple(meta["pk"]) for name, meta in sync_targets().items()}
    targets.update(
        {
            "ana_stock_concept_map": ("ts_code", "concept_code"),
            "ana_concept_day": ("concept_code", "trade_date"),
        }
    )
    return targets


def build_upsert_sql(table_name: str, columns: list[str], conflict_keys: list[str]) -> str:
    insert_cols = ", ".join(columns)
    placeholders = ", ".join(["%s"] * len(columns))
    conflict_expr = ", ".join(conflict_keys)
    update_cols = [c for c in columns if c not in conflict_keys]
    update_expr = ", ".join(f"{col} = EXCLUDED.{col}" for col in update_cols)
    return (
        f"INSERT INTO {table_name} ({insert_cols}) VALUES ({placeholders}) "
        f"ON CONFLICT ({conflict_expr}) DO UPDATE SET {update_expr}"
    )


def build_sync_state_upsert_sql() -> str:
    return """
INSERT INTO sync_job_state (
    job_name,
    target_table,
    target_key,
    last_success_cursor,
    last_success_at,
    status,
    error_message,
    updated_at
) VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
ON CONFLICT (job_name, target_table, target_key)
DO UPDATE SET
    last_success_cursor = EXCLUDED.last_success_cursor,
    last_success_at = EXCLUDED.last_success_at,
    status = EXCLUDED.status,
    error_message = EXCLUDED.error_message,
    updated_at = NOW()
""".strip()


def requests_per_min_to_sleep_seconds(requests_per_min: int) -> float:
    return round(60.0 / float(requests_per_min), 6)


def is_tushare_rate_limit_error(exc: Exception) -> bool:
    message = str(exc)
    return "每分钟最多访问该接口" in message or "分钟最多访问" in message


def is_tushare_retryable_error(exc: Exception) -> bool:
    message = str(exc)
    return (
        is_tushare_rate_limit_error(exc)
        or "服务器内部错误，请稍后重试" in message
        or "HTTPConnectionPool(" in message
        or "NameResolutionError" in message
        or "Failed to resolve" in message
        or "Max retries exceeded" in message
    )


def call_with_rate_limit_retry(
    func: Callable[[], object],
    max_attempts: int = 3,
    sleep_seconds: float = 15.0,
    sleeper: Callable[[float], None] | None = None,
):
    sleeper = sleeper or time.sleep
    for attempt in range(1, max_attempts + 1):
        try:
            return func()
        except Exception as exc:
            if not is_tushare_retryable_error(exc) or attempt == max_attempts:
                raise
            sleeper(sleep_seconds)


def format_state_cursor(ts_code: str | None, cursor_date: str) -> str:
    if ts_code:
        return f"{ts_code}|{cursor_date}"
    return cursor_date


def parse_state_cursor(cursor: str | None) -> tuple[str | None, str | None]:
    if not cursor:
        return None, None
    if "|" in cursor:
        ts_code, cursor_date = cursor.split("|", 1)
        return ts_code, cursor_date
    return None, cursor


def compute_resume_start_date(cursor_date: str | None, overlap_days: int = 7) -> str | None:
    if not cursor_date:
        return None
    dt = datetime.strptime(cursor_date, "%Y-%m-%d")
    return (dt - timedelta(days=overlap_days)).strftime("%Y%m%d")


def build_stock_resume_plan(
    stock_codes: list[str],
    requested_start_date: str,
    last_cursor: str | None,
    overlap_days: int = 7,
) -> list[tuple[str, str]]:
    cursor_code, cursor_date = parse_state_cursor(last_cursor)
    if not cursor_code or cursor_code not in stock_codes:
        return [(code, requested_start_date) for code in stock_codes]

    start_idx = stock_codes.index(cursor_code)
    resumed_start = compute_resume_start_date(cursor_date, overlap_days=overlap_days) or requested_start_date
    plan: list[tuple[str, str]] = []
    for idx, code in enumerate(stock_codes[start_idx:], start=start_idx):
        plan.append((code, resumed_start if idx == start_idx else requested_start_date))
    return plan


def build_code_resume_plan(codes: list[str], last_cursor: str | None) -> list[str]:
    cursor_code, _ = parse_state_cursor(last_cursor)
    if not cursor_code or cursor_code not in codes:
        return codes
    start_idx = codes.index(cursor_code) + 1
    return codes[start_idx:]


def load_case_stock_names(workbook_path: str | Path, sheet_name: str = DEFAULT_CASE_SHEET_NAME) -> list[str]:
    workbook = load_workbook(filename=workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        stock_name_idx = headers.index("标的名称")
        ordered_names: list[str] = []
        seen: set[str] = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            stock_name = row[stock_name_idx]
            if not stock_name or stock_name in seen:
                continue
            seen.add(stock_name)
            ordered_names.append(stock_name)
        return ordered_names
    finally:
        workbook.close()


def load_stock_codes_from_csv(csv_path: str | Path, code_column: str = "股票代码") -> list[str]:
    df = pd.read_csv(csv_path)
    ordered_codes: list[str] = []
    seen: set[str] = set()
    for raw_code in df[code_column].dropna().tolist():
        ts_code = str(raw_code).strip()
        if not ts_code or ts_code in seen:
            continue
        seen.add(ts_code)
        ordered_codes.append(ts_code)
    return ordered_codes


def resolve_case_stock_codes(stock_names: list[str], stock_basic_df: pd.DataFrame) -> list[str]:
    name_to_code = {
        row["name"]: row["ts_code"]
        for row in stock_basic_df[["ts_code", "name"]].dropna().to_dict("records")
    }
    stock_codes: list[str] = []
    missing_names: list[str] = []
    for stock_name in stock_names:
        ts_code = name_to_code.get(stock_name)
        if not ts_code:
            missing_names.append(stock_name)
            continue
        stock_codes.append(ts_code)
    if missing_names:
        missing_text = ", ".join(missing_names)
        raise ValueError(f"未找到股票代码: {missing_text}")
    return sorted(stock_codes)


def build_case_stock_sync_config(
    stock_names: list[str],
    stock_basic_df: pd.DataFrame,
    start_date: str,
    job_name: str = DEFAULT_CASE_STOCK_BUNDLE_JOB_NAME,
    target_key: str = DEFAULT_CASE_STOCK_BUNDLE_TARGET_KEY,
    exclude_names: list[str] | None = None,
) -> dict[str, object]:
    exclude_set = set(exclude_names or [])
    filtered_stock_names = [stock_name for stock_name in stock_names if stock_name not in exclude_set]
    return {
        "job_name": job_name,
        "target_key": target_key,
        "start_date": start_date,
        "stock_names": filtered_stock_names,
        "stock_codes": resolve_case_stock_codes(filtered_stock_names, stock_basic_df),
    }


def build_stock_bundle_sync_config(
    stock_codes: list[str],
    start_date: str,
    job_name: str,
    target_key: str,
) -> dict[str, object]:
    deduped_codes = sorted({str(code).strip() for code in stock_codes if str(code).strip()})
    return {
        "job_name": job_name,
        "target_key": target_key,
        "start_date": start_date,
        "stock_codes": deduped_codes,
    }


def build_trade_date_resume_plan(
    trade_dates: list[str],
    last_cursor: str | None,
    overlap_days: int = 7,
) -> list[str]:
    if not last_cursor:
        return trade_dates
    _, cursor_date = parse_state_cursor(last_cursor)
    resumed_start = compute_resume_start_date(cursor_date, overlap_days=overlap_days)
    if not resumed_start:
        return trade_dates
    return [trade_date for trade_date in trade_dates if trade_date >= resumed_start]


def run_case_stock_bundle_sync(
    sync_config: dict[str, object],
    last_cursor: str | None,
    fetch_stock_bundle: Callable[[str, str, str], dict[str, pd.DataFrame]],
    persist_frames: Callable[[str, dict[str, pd.DataFrame]], None],
    persist_sync_state: Callable[[dict[str, object]], None],
    end_date: str,
    overlap_days: int = 7,
    sleep_seconds: float = 0.0,
    sleeper: Callable[[float], None] | None = None,
) -> None:
    sleeper = sleeper or time.sleep
    stock_codes = list(sync_config["stock_codes"])
    latest_success_cursor: str | None = last_cursor
    plan = build_stock_resume_plan(
        stock_codes=stock_codes,
        requested_start_date=str(sync_config["start_date"]),
        last_cursor=last_cursor,
        overlap_days=overlap_days,
    )
    for ts_code, start_date in plan:
        try:
            frames = fetch_stock_bundle(ts_code, start_date, end_date)
            persist_frames(ts_code, frames)
            persist_sync_state(
                {
                    "job_name": sync_config["job_name"],
                    "target_table": "raw_stock_daily_qfq",
                    "target_key": sync_config["target_key"],
                    "last_success_cursor": format_state_cursor(ts_code, _normalize_cursor_date(end_date)),
                    "last_success_at": datetime.now(),
                    "status": "success",
                    "error_message": None,
                }
            )
            latest_success_cursor = format_state_cursor(ts_code, _normalize_cursor_date(end_date))
        except Exception as exc:
            persist_sync_state(
                {
                    "job_name": sync_config["job_name"],
                    "target_table": "raw_stock_daily_qfq",
                    "target_key": sync_config["target_key"],
                    "last_success_cursor": latest_success_cursor,
                    "last_success_at": datetime.now(),
                    "status": "failed",
                    "error_message": str(exc),
                }
            )
            raise
        if sleep_seconds > 0:
            sleeper(sleep_seconds)


def run_stock_concept_bundle_sync(
    sync_config: dict[str, object],
    last_cursor: str | None,
    fetch_concept_bundle: Callable[[str, str, str], dict[str, pd.DataFrame]],
    persist_frames: Callable[[str, dict[str, pd.DataFrame]], None],
    persist_sync_state: Callable[[dict[str, object]], None],
    end_date: str,
    overlap_days: int = 7,
    sleep_seconds: float = 0.0,
    sleeper: Callable[[float], None] | None = None,
) -> None:
    sleeper = sleeper or time.sleep
    concept_codes = list(sync_config["concept_codes"])
    latest_success_cursor: str | None = last_cursor
    plan = build_stock_resume_plan(
        stock_codes=concept_codes,
        requested_start_date=str(sync_config["start_date"]),
        last_cursor=last_cursor,
        overlap_days=overlap_days,
    )
    for concept_code, start_date in plan:
        try:
            frames = fetch_concept_bundle(concept_code, start_date, end_date)
            persist_frames(concept_code, frames)
            latest_success_cursor = format_state_cursor(concept_code, _normalize_cursor_date(end_date))
            persist_sync_state(
                {
                    "job_name": sync_config["job_name"],
                    "target_table": "raw_ths_member",
                    "target_key": sync_config["target_key"],
                    "last_success_cursor": latest_success_cursor,
                    "last_success_at": datetime.now(),
                    "status": "success",
                    "error_message": None,
                }
            )
        except Exception as exc:
            persist_sync_state(
                {
                    "job_name": sync_config["job_name"],
                    "target_table": "raw_ths_member",
                    "target_key": sync_config["target_key"],
                    "last_success_cursor": latest_success_cursor,
                    "last_success_at": datetime.now(),
                    "status": "failed",
                    "error_message": str(exc),
                }
            )
            raise
        if sleep_seconds > 0:
            sleeper(sleep_seconds)


def run_concept_daily_sync(
    sync_config: dict[str, object],
    last_cursor: str | None,
    fetch_concept_daily_bundle: Callable[[str, str, str], dict[str, pd.DataFrame]],
    persist_frames: Callable[[str, dict[str, pd.DataFrame]], None],
    persist_sync_state: Callable[[dict[str, object]], None],
    end_date: str,
    overlap_days: int = 7,
    sleep_seconds: float = 0.0,
    sleeper: Callable[[float], None] | None = None,
) -> None:
    sleeper = sleeper or time.sleep
    concept_codes = list(sync_config["concept_codes"])
    latest_success_cursor: str | None = last_cursor
    plan = build_stock_resume_plan(
        stock_codes=concept_codes,
        requested_start_date=str(sync_config["start_date"]),
        last_cursor=last_cursor,
        overlap_days=overlap_days,
    )
    for concept_code, start_date in plan:
        try:
            frames = fetch_concept_daily_bundle(concept_code, start_date, end_date)
            persist_frames(concept_code, frames)
            latest_success_cursor = format_state_cursor(concept_code, _normalize_cursor_date(end_date))
            persist_sync_state(
                {
                    "job_name": sync_config["job_name"],
                    "target_table": "raw_ths_concept_daily",
                    "target_key": sync_config["target_key"],
                    "last_success_cursor": latest_success_cursor,
                    "last_success_at": datetime.now(),
                    "status": "success",
                    "error_message": None,
                }
            )
        except Exception as exc:
            persist_sync_state(
                {
                    "job_name": sync_config["job_name"],
                    "target_table": "raw_ths_concept_daily",
                    "target_key": sync_config["target_key"],
                    "last_success_cursor": latest_success_cursor,
                    "last_success_at": datetime.now(),
                    "status": "failed",
                    "error_message": str(exc),
                }
            )
            raise
        if sleep_seconds > 0:
            sleeper(sleep_seconds)


def run_concept_member_sync(
    sync_config: dict[str, object],
    last_cursor: str | None,
    fetch_concept_member_bundle: Callable[[str], dict[str, pd.DataFrame]],
    persist_frames: Callable[[str, dict[str, pd.DataFrame]], None],
    persist_sync_state: Callable[[dict[str, object]], None],
    end_date: str,
    sleep_seconds: float = 0.0,
    sleeper: Callable[[float], None] | None = None,
) -> None:
    sleeper = sleeper or time.sleep
    concept_codes = list(sync_config["concept_codes"])
    latest_success_cursor: str | None = last_cursor
    plan = build_code_resume_plan(concept_codes, last_cursor)
    for concept_code in plan:
        try:
            frames = fetch_concept_member_bundle(concept_code)
            persist_frames(concept_code, frames)
            latest_success_cursor = format_state_cursor(concept_code, _normalize_cursor_date(end_date))
            persist_sync_state(
                {
                    "job_name": sync_config["job_name"],
                    "target_table": "raw_ths_member",
                    "target_key": sync_config["target_key"],
                    "last_success_cursor": latest_success_cursor,
                    "last_success_at": datetime.now(),
                    "status": "success",
                    "error_message": None,
                }
            )
        except Exception as exc:
            persist_sync_state(
                {
                    "job_name": sync_config["job_name"],
                    "target_table": "raw_ths_member",
                    "target_key": sync_config["target_key"],
                    "last_success_cursor": latest_success_cursor,
                    "last_success_at": datetime.now(),
                    "status": "failed",
                    "error_message": str(exc),
                }
            )
            raise
        if sleep_seconds > 0:
            sleeper(sleep_seconds)


def _normalize_cursor_date(date_text: str) -> str:
    if "-" in date_text:
        return date_text
    return datetime.strptime(date_text, "%Y%m%d").strftime("%Y-%m-%d")


def create_tushare_client(token: str, http_url: str = DEFAULT_TUSHARE_HTTP_URL):
    ts.set_token(token)
    pro = ts.pro_api(token)
    pro._DataApi__token = token
    pro._DataApi__http_url = http_url
    return pro


def list_open_trade_dates(pro, start_date: str, end_date: str) -> list[str]:
    trade_cal_df = pro.trade_cal(exchange="", start_date=start_date, end_date=end_date, is_open=1)
    if trade_cal_df is None or trade_cal_df.empty:
        return []
    return sorted(trade_cal_df["cal_date"].astype(str).dropna().tolist())


def load_latest_adj_factor_snapshot(
    pro,
    anchor_trade_date: str,
    per_request_sleep_seconds: float = 0.0,
    sleeper: Callable[[float], None] | None = None,
) -> dict[str, float]:
    sleeper = sleeper or time.sleep
    logger.info("开始抓取全市场最新复权因子快照: trade_date=%s", anchor_trade_date)
    adj_factor_df = call_with_rate_limit_retry(lambda: pro.adj_factor(trade_date=anchor_trade_date))
    if per_request_sleep_seconds > 0:
        sleeper(per_request_sleep_seconds)
    if adj_factor_df is None or adj_factor_df.empty:
        raise ValueError(f"未获取到 {anchor_trade_date} 的全市场复权因子快照")
    result = {
        str(row.ts_code): float(row.adj_factor)
        for row in adj_factor_df[["ts_code", "adj_factor"]].dropna().itertuples(index=False)
    }
    logger.info("全市场最新复权因子快照抓取完成: trade_date=%s stocks=%s", anchor_trade_date, len(result))
    return result


@contextmanager
def requests_sessions_without_proxy():
    original_init = requests.sessions.Session.__init__

    def patched_init(self, *args, **kwargs):
        original_init(self, *args, **kwargs)
        self.trust_env = False

    requests.sessions.Session.__init__ = patched_init
    try:
        yield
    finally:
        requests.sessions.Session.__init__ = original_init


def fetch_case_stock_bundle(
    pro,
    ts_code: str,
    start_date: str,
    end_date: str,
    per_request_sleep_seconds: float = 0.0,
    sleeper: Callable[[float], None] | None = None,
    ak_client=None,
) -> dict[str, pd.DataFrame]:
    sleeper = sleeper or time.sleep
    ak_client = ak_client or ak

    def _call(func: Callable[[], object]):
        result = call_with_rate_limit_retry(func)
        if per_request_sleep_seconds > 0:
            sleeper(per_request_sleep_seconds)
        return result
    logger.info(
        "开始抓取股票包[Tushare]: ts_code=%s start_date=%s end_date=%s",
        ts_code,
        start_date,
        end_date,
    )
    try:
        daily_df = _call(lambda: pro.daily(ts_code=ts_code, start_date=start_date, end_date=end_date))
        adj_factor_df = _call(lambda: pro.adj_factor(ts_code=ts_code, start_date=start_date, end_date=end_date))
        daily_basic_df = _call(lambda: pro.daily_basic(ts_code=ts_code, start_date=start_date, end_date=end_date))
        moneyflow_df = _call(lambda: pro.moneyflow(ts_code=ts_code, start_date=start_date, end_date=end_date))
        limit_list_df = _call(lambda: pro.limit_list_d(ts_code=ts_code, start_date=start_date, end_date=end_date))
        result = {
            "raw_stock_daily_qfq": build_qfq_daily_frame(daily_df, adj_factor_df),
            "raw_daily_basic": normalize_daily_basic(daily_basic_df) if daily_basic_df is not None and not daily_basic_df.empty else pd.DataFrame(),
            "raw_moneyflow": normalize_moneyflow(moneyflow_df) if moneyflow_df is not None and not moneyflow_df.empty else pd.DataFrame(),
            "raw_limit_list_d": normalize_limit_list_d(limit_list_df) if limit_list_df is not None and not limit_list_df.empty else pd.DataFrame(),
        }
        logger.info(
            "股票包抓取完成[Tushare]: ts_code=%s rows=%s",
            ts_code,
            {name: len(frame) for name, frame in result.items()},
        )
        return result
    except Exception as exc:
        logger.warning(
            "Tushare 抓取股票包失败，回退到 Akshare: ts_code=%s start_date=%s end_date=%s error=%s",
            ts_code,
            start_date,
            end_date,
            exc,
        )
        try:
            result = fetch_case_stock_bundle_from_akshare(
                ts_code=ts_code,
                start_date=start_date,
                end_date=end_date,
                ak_client=ak_client,
            )
            logger.info(
                "股票包抓取完成[Akshare]: ts_code=%s rows=%s",
                ts_code,
                {name: len(frame) for name, frame in result.items()},
            )
            return result
        except Exception as ak_exc:
            logger.error(
                "Akshare fallback 也失败: ts_code=%s start_date=%s end_date=%s",
                ts_code,
                start_date,
                end_date,
                exc_info=True,
            )
            raise RuntimeError(f"Tushare 与 Akshare 均失败: {exc}; {ak_exc}") from ak_exc


def load_sync_state(conn, job_name: str, target_table: str, target_key: str) -> dict[str, object] | None:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT job_name, target_table, target_key, last_success_cursor, last_success_at, status, error_message
            FROM sync_job_state
            WHERE job_name = %s AND target_table = %s AND target_key = %s
            """,
            (job_name, target_table, target_key),
        )
        row = cur.fetchone()
    if not row:
        return None
    return {
        "job_name": row[0],
        "target_table": row[1],
        "target_key": row[2],
        "last_success_cursor": row[3],
        "last_success_at": row[4],
        "status": row[5],
        "error_message": row[6],
    }


def persist_sync_state_row(conn, payload: dict[str, object]) -> None:
    with conn.cursor() as cur:
        cur.execute(
            build_sync_state_upsert_sql(),
            (
                payload["job_name"],
                payload["target_table"],
                payload["target_key"],
                payload["last_success_cursor"],
                payload["last_success_at"],
                payload["status"],
                payload["error_message"],
            ),
        )
    conn.commit()


def persist_frames_to_db(conn, ts_code: str, frames: dict[str, pd.DataFrame]) -> None:
    del ts_code
    pk_map = table_primary_keys()
    for table_name, frame in frames.items():
        if frame.empty:
            continue
        columns = list(frame.columns)
        conflict_keys = list(pk_map[table_name])
        sql = build_upsert_sql(table_name, columns, conflict_keys)
        rows = [tuple(row[col] for col in columns) for row in frame.to_dict("records")]
        with conn.cursor() as cur:
            cur.executemany(sql, rows)
    conn.commit()


def build_market_qfq_daily_frame(
    daily_df: pd.DataFrame,
    adj_factor_df: pd.DataFrame,
    latest_adj_factor_map: dict[str, float],
) -> pd.DataFrame:
    if daily_df is None or daily_df.empty:
        return _empty_stock_daily_qfq_frame()

    required_columns = {"ts_code", "trade_date", "adj_factor"}
    if adj_factor_df is None or adj_factor_df.empty or not required_columns.issubset(set(adj_factor_df.columns)):
        logger.warning("全市场切片缺少 adj_factor，回退为原始日线价格: rows=%s", len(daily_df))
        return normalize_stock_daily_qfq(daily_df)

    merged = daily_df.merge(
        adj_factor_df[["ts_code", "trade_date", "adj_factor"]],
        on=["ts_code", "trade_date"],
        how="left",
    )
    latest_factor = merged["ts_code"].map(latest_adj_factor_map)
    latest_factor = latest_factor.where(latest_factor.notna(), merged["adj_factor"])
    factor_scale = pd.Series(1.0, index=merged.index, dtype="float64")
    valid_mask = merged["adj_factor"].notna() & latest_factor.notna() & (latest_factor != 0)
    factor_scale.loc[valid_mask] = merged.loc[valid_mask, "adj_factor"] / latest_factor.loc[valid_mask]

    merged["open_qfq"] = merged["open"] * factor_scale
    merged["high_qfq"] = merged["high"] * factor_scale
    merged["low_qfq"] = merged["low"] * factor_scale
    merged["close_qfq"] = merged["close"] * factor_scale

    return merged[
        [
            "ts_code",
            "trade_date",
            "open_qfq",
            "high_qfq",
            "low_qfq",
            "close_qfq",
            "pct_chg",
            "vol",
            "amount",
        ]
    ].copy()


def fetch_market_trade_date_bundle(
    pro,
    trade_date: str,
    latest_adj_factor_map: dict[str, float],
    per_request_sleep_seconds: float = 0.0,
    sleeper: Callable[[float], None] | None = None,
    max_workers: int = DEFAULT_MAX_WORKERS,
) -> dict[str, pd.DataFrame]:
    sleeper = sleeper or time.sleep

    def _call(func: Callable[[], object]):
        result = call_with_rate_limit_retry(func)
        if per_request_sleep_seconds > 0:
            sleeper(per_request_sleep_seconds)
        return result

    logger.info("开始抓取全市场日切片: trade_date=%s", trade_date)
    call_specs: OrderedDict[str, Callable[[], object]] = OrderedDict(
        [
            ("daily", lambda: pro.daily(trade_date=trade_date)),
            ("adj_factor", lambda: pro.adj_factor(trade_date=trade_date)),
            ("daily_basic", lambda: pro.daily_basic(trade_date=trade_date)),
            ("moneyflow", lambda: pro.moneyflow(trade_date=trade_date)),
            ("limit_list_d", lambda: pro.limit_list_d(trade_date=trade_date)),
        ]
    )

    results: dict[str, object] = {}
    workers = max(1, min(max_workers, len(call_specs)))
    if workers == 1:
        for name, func in call_specs.items():
            results[name] = _call(func)
    else:
        with ThreadPoolExecutor(max_workers=workers) as executor:
            future_map = {
                executor.submit(_call, func): name
                for name, func in call_specs.items()
            }
            for future, name in future_map.items():
                results[name] = future.result()

    daily_df = results["daily"]
    adj_factor_df = results["adj_factor"]
    daily_basic_df = results["daily_basic"]
    moneyflow_df = results["moneyflow"]
    limit_list_df = results["limit_list_d"]
    bundle = {
        "raw_stock_daily_qfq": build_market_qfq_daily_frame(daily_df, adj_factor_df, latest_adj_factor_map),
        "raw_daily_basic": normalize_daily_basic(daily_basic_df) if daily_basic_df is not None and not daily_basic_df.empty else _empty_daily_basic_frame(),
        "raw_moneyflow": normalize_moneyflow(moneyflow_df) if moneyflow_df is not None and not moneyflow_df.empty else _empty_moneyflow_frame(),
        "raw_limit_list_d": normalize_limit_list_d(limit_list_df) if limit_list_df is not None and not limit_list_df.empty else _empty_limit_list_frame(),
    }
    logger.info(
        "全市场日切片抓取完成: trade_date=%s rows=%s",
        trade_date,
        {name: len(frame) for name, frame in bundle.items()},
    )
    return bundle


def run_market_trade_date_sync(
    sync_config: dict[str, object],
    last_cursor: str | None,
    fetch_trade_date_bundle: Callable[[str], dict[str, pd.DataFrame]],
    persist_frames: Callable[[str, dict[str, pd.DataFrame]], None],
    persist_sync_state: Callable[[dict[str, object]], None],
    overlap_days: int = 7,
) -> None:
    trade_dates = list(sync_config["trade_dates"])
    latest_success_cursor: str | None = last_cursor
    plan = build_trade_date_resume_plan(trade_dates=trade_dates, last_cursor=last_cursor, overlap_days=overlap_days)
    for trade_date in plan:
        try:
            frames = fetch_trade_date_bundle(trade_date)
            persist_frames(trade_date, frames)
            latest_success_cursor = _normalize_cursor_date(trade_date)
            persist_sync_state(
                {
                    "job_name": sync_config["job_name"],
                    "target_table": "raw_stock_daily_qfq",
                    "target_key": sync_config["target_key"],
                    "last_success_cursor": latest_success_cursor,
                    "last_success_at": datetime.now(),
                    "status": "success",
                    "error_message": None,
                }
            )
        except Exception as exc:
            persist_sync_state(
                {
                    "job_name": sync_config["job_name"],
                    "target_table": "raw_stock_daily_qfq",
                    "target_key": sync_config["target_key"],
                    "last_success_cursor": latest_success_cursor,
                    "last_success_at": datetime.now(),
                    "status": "failed",
                    "error_message": str(exc),
                }
            )
            raise


def fetch_stock_concept_bundle(
    pro,
    concept_code: str,
    concept_name: str,
    stock_codes: list[str],
    start_date: str,
    end_date: str,
    mapping_asof_date: str,
    per_request_sleep_seconds: float = 0.0,
    sleeper: Callable[[float], None] | None = None,
) -> dict[str, pd.DataFrame]:
    sleeper = sleeper or time.sleep
    stock_set = set(stock_codes)

    def _call(func: Callable[[], object]):
        result = call_with_rate_limit_retry(func)
        if per_request_sleep_seconds > 0:
            sleeper(per_request_sleep_seconds)
        return result

    member_df = _call(lambda: pro.ths_member(ts_code=concept_code))
    if member_df is None or member_df.empty:
        empty_member = pd.DataFrame(columns=["ts_code", "con_code", "con_name", "mapping_asof_date"])
        empty_concepts = pd.DataFrame(columns=["concept_code", "trade_date", "concept_name", "close", "pct_change", "vol", "turnover_rate"])
        empty_map = pd.DataFrame(columns=["ts_code", "concept_code", "concept_name", "mapping_asof_date", "map_source", "updated_at"])
        return {
            "raw_ths_member": empty_member,
            "raw_ths_concept_daily": pd.DataFrame(columns=["ts_code", "trade_date", "concept_name", "open", "high", "low", "close", "pct_change", "vol", "turnover_rate"]),
            "ana_stock_concept_map": empty_map,
            "ana_concept_day": empty_concepts,
        }

    hits = member_df[member_df["con_code"].isin(stock_set)].copy()
    if hits.empty:
        empty_member = pd.DataFrame(columns=["ts_code", "con_code", "con_name", "mapping_asof_date"])
        empty_concepts = pd.DataFrame(columns=["concept_code", "trade_date", "concept_name", "close", "pct_change", "vol", "turnover_rate"])
        empty_map = pd.DataFrame(columns=["ts_code", "concept_code", "concept_name", "mapping_asof_date", "map_source", "updated_at"])
        return {
            "raw_ths_member": empty_member,
            "raw_ths_concept_daily": pd.DataFrame(columns=["ts_code", "trade_date", "concept_name", "open", "high", "low", "close", "pct_change", "vol", "turnover_rate"]),
            "ana_stock_concept_map": empty_map,
            "ana_concept_day": empty_concepts,
        }

    raw_member = normalize_ths_member(hits, mapping_asof_date)
    concept_meta = pd.DataFrame([{"concept_code": concept_code, "concept_name": concept_name}])
    concept_daily_df = _call(lambda: pro.ths_daily(ts_code=concept_code, start_date=start_date, end_date=end_date))
    raw_concept_daily = (
        normalize_ths_concept_daily(concept_daily_df, concept_name)
        if concept_daily_df is not None and not concept_daily_df.empty
        else pd.DataFrame(columns=["ts_code", "trade_date", "concept_name", "open", "high", "low", "close", "pct_change", "vol", "turnover_rate"])
    )
    return {
        "raw_ths_member": raw_member,
        "raw_ths_concept_daily": raw_concept_daily,
        "ana_stock_concept_map": normalize_ana_stock_concept_map(raw_member, concept_meta),
        "ana_concept_day": normalize_ana_concept_day(raw_concept_daily),
    }


def fetch_concept_daily_bundle(
    pro,
    concept_code: str,
    concept_name: str,
    start_date: str,
    end_date: str,
    per_request_sleep_seconds: float = 0.0,
    sleeper: Callable[[float], None] | None = None,
) -> dict[str, pd.DataFrame]:
    sleeper = sleeper or time.sleep

    def _call(func: Callable[[], object]):
        result = call_with_rate_limit_retry(func)
        if per_request_sleep_seconds > 0:
            sleeper(per_request_sleep_seconds)
        return result

    logger.info(
        "开始抓取概念指数包[Tushare]: concept_code=%s concept_name=%s start_date=%s end_date=%s",
        concept_code,
        concept_name,
        start_date,
        end_date,
    )
    concept_daily_df = _call(lambda: pro.ths_daily(ts_code=concept_code, start_date=start_date, end_date=end_date))
    raw_concept_daily = (
        normalize_ths_concept_daily(concept_daily_df, concept_name)
        if concept_daily_df is not None and not concept_daily_df.empty
        else pd.DataFrame(columns=["ts_code", "trade_date", "concept_name", "open", "high", "low", "close", "pct_change", "vol", "turnover_rate"])
    )
    frames = {
        "raw_ths_concept_daily": raw_concept_daily,
        "ana_concept_day": normalize_ana_concept_day(raw_concept_daily),
    }
    logger.info(
        "概念指数包抓取完成[Tushare]: concept_code=%s rows=%s",
        concept_code,
        {name: len(frame) for name, frame in frames.items()},
    )
    return frames


def fetch_case_stock_concept_bundle_from_tushare(
    ts_code: str,
    start_date: str,
    end_date: str,
    token: str,
    http_url: str = DEFAULT_TUSHARE_HTTP_URL,
    pro=None,
    per_request_sleep_seconds: float = 0.0,
    sleeper: Callable[[float], None] | None = None,
) -> dict[str, pd.DataFrame]:
    sleeper = sleeper or time.sleep
    pro = pro or create_tushare_client(token=token, http_url=http_url)
    mapping_asof_date = end_date

    def _call(func: Callable[[], object]):
        result = call_with_rate_limit_retry(func)
        if per_request_sleep_seconds > 0:
            sleeper(per_request_sleep_seconds)
        return result

    logger.info(
        "开始抓取股票概念包[Tushare]: ts_code=%s start_date=%s end_date=%s http_url=%s",
        ts_code,
        start_date,
        end_date,
        http_url,
    )
    index_df = _call(lambda: pro.ths_index())
    member_frame, concept_frame = build_theme_concept_matches(
        index_df=index_df,
        stock_codes=[ts_code],
        fetch_members=lambda concept_code: _call(lambda: pro.ths_member(ts_code=concept_code)),
        mapping_asof_date=mapping_asof_date,
    )
    if member_frame.empty or concept_frame.empty:
        empty_member = pd.DataFrame(columns=["ts_code", "con_code", "con_name", "mapping_asof_date"])
        empty_concepts = pd.DataFrame(columns=["concept_code", "trade_date", "concept_name", "close", "pct_change", "vol", "turnover_rate"])
        empty_map = pd.DataFrame(columns=["ts_code", "concept_code", "concept_name", "mapping_asof_date", "map_source", "updated_at"])
        return {
            "raw_ths_member": empty_member,
            "raw_ths_concept_daily": pd.DataFrame(columns=["ts_code", "trade_date", "concept_name", "open", "high", "low", "close", "pct_change", "vol", "turnover_rate"]),
            "ana_stock_concept_map": empty_map,
            "ana_concept_day": empty_concepts,
        }

    raw_daily_frames: list[pd.DataFrame] = []
    for row in concept_frame.itertuples(index=False):
        concept_daily_df = _call(lambda: pro.ths_daily(ts_code=row.concept_code, start_date=start_date, end_date=end_date))
        if concept_daily_df is None or concept_daily_df.empty:
            continue
        raw_daily_frames.append(normalize_ths_concept_daily(concept_daily_df, row.concept_name))

    raw_concept_daily = (
        pd.concat(raw_daily_frames, ignore_index=True).drop_duplicates(subset=["ts_code", "trade_date"]).reset_index(drop=True)
        if raw_daily_frames
        else pd.DataFrame(columns=["ts_code", "trade_date", "concept_name", "open", "high", "low", "close", "pct_change", "vol", "turnover_rate"])
    )
    return {
        "raw_ths_member": member_frame,
        "raw_ths_concept_daily": raw_concept_daily,
        "ana_stock_concept_map": normalize_ana_stock_concept_map(member_frame, concept_frame),
        "ana_concept_day": normalize_ana_concept_day(raw_concept_daily),
    }


def fetch_concept_member_bundle(
    pro,
    concept_code: str,
    concept_name: str,
    mapping_asof_date: str,
    per_request_sleep_seconds: float = 0.0,
    sleeper: Callable[[float], None] | None = None,
) -> dict[str, pd.DataFrame]:
    sleeper = sleeper or time.sleep

    def _call(func: Callable[[], object]):
        result = call_with_rate_limit_retry(func)
        if per_request_sleep_seconds > 0:
            sleeper(per_request_sleep_seconds)
        return result

    logger.info(
        "开始抓取概念成员包[Tushare]: concept_code=%s concept_name=%s mapping_asof_date=%s",
        concept_code,
        concept_name,
        mapping_asof_date,
    )
    member_df = _call(lambda: pro.ths_member(ts_code=concept_code))
    if member_df is None or member_df.empty:
        empty_member = pd.DataFrame(columns=["ts_code", "con_code", "con_name", "mapping_asof_date"])
        empty_map = pd.DataFrame(columns=["ts_code", "concept_code", "concept_name", "mapping_asof_date", "map_source", "updated_at"])
        return {
            "raw_ths_member": empty_member,
            "ana_stock_concept_map": empty_map,
        }

    raw_member = normalize_ths_member(member_df, mapping_asof_date)
    concept_meta = pd.DataFrame([{"concept_code": concept_code, "concept_name": concept_name}])
    frames = {
        "raw_ths_member": raw_member,
        "ana_stock_concept_map": normalize_ana_stock_concept_map(raw_member, concept_meta),
    }
    logger.info(
        "概念成员包抓取完成[Tushare]: concept_code=%s rows=%s",
        concept_code,
        {name: len(frame) for name, frame in frames.items()},
    )
    return frames


def sync_stock_file_concepts_to_db(
    stock_file_path: str | Path,
    start_date: str,
    end_date: str,
    token: str,
    job_name: str,
    target_key: str,
    db_dsn: str = DEFAULT_DB_DSN,
    http_url: str = DEFAULT_TUSHARE_HTTP_URL,
    requests_per_min: int = 20,
    overlap_days: int = 7,
    code_column: str = "股票代码",
    per_request_sleep_seconds: float | None = None,
) -> dict[str, object]:
    pro = create_tushare_client(token=token, http_url=http_url)
    stock_codes = load_stock_codes_from_csv(stock_file_path, code_column=code_column)
    index_df = filter_ths_index_targets(pro.ths_index())
    sync_config = {
        "job_name": job_name,
        "target_key": target_key,
        "start_date": start_date,
        "stock_codes": stock_codes,
        "concept_codes": index_df["ts_code"].astype(str).tolist(),
    }
    concept_name_map = dict(zip(index_df["ts_code"], index_df["name"]))
    sleep_seconds = requests_per_min_to_sleep_seconds(requests_per_min)
    per_request_sleep_seconds = per_request_sleep_seconds if per_request_sleep_seconds is not None else sleep_seconds
    mapping_asof_date = datetime.now().strftime("%Y%m%d")

    with psycopg.connect(db_dsn) as conn:
        state = load_sync_state(
            conn,
            job_name=str(sync_config["job_name"]),
            target_table="raw_ths_member",
            target_key=str(sync_config["target_key"]),
        )
        last_cursor = state["last_success_cursor"] if state else None

        def _fetch(concept_code: str, planned_start_date: str, planned_end_date: str) -> dict[str, pd.DataFrame]:
            return fetch_stock_concept_bundle(
                pro=pro,
                concept_code=concept_code,
                concept_name=concept_name_map[concept_code],
                stock_codes=stock_codes,
                start_date=planned_start_date,
                end_date=planned_end_date,
                mapping_asof_date=mapping_asof_date,
                per_request_sleep_seconds=per_request_sleep_seconds,
            )

        def _persist_frames(concept_code: str, frames: dict[str, pd.DataFrame]) -> None:
            persist_frames_to_db(conn, concept_code, frames)

        def _persist_state(payload: dict[str, object]) -> None:
            persist_sync_state_row(conn, payload)

        run_stock_concept_bundle_sync(
            sync_config=sync_config,
            last_cursor=last_cursor,
            fetch_concept_bundle=_fetch,
            persist_frames=_persist_frames,
            persist_sync_state=_persist_state,
            end_date=end_date,
            overlap_days=overlap_days,
            sleep_seconds=0.0,
        )
    return sync_config


def sync_all_concepts_to_db(
    start_date: str,
    end_date: str,
    token: str,
    job_name: str = DEFAULT_ALL_CONCEPTS_JOB_NAME,
    target_key: str = DEFAULT_ALL_CONCEPTS_TARGET_KEY,
    db_dsn: str = DEFAULT_DB_DSN,
    http_url: str = DEFAULT_TUSHARE_HTTP_URL,
    requests_per_min: int = DEFAULT_FEATURE_REQUESTS_PER_MIN,
    overlap_days: int = 7,
    per_request_sleep_seconds: float | None = None,
) -> dict[str, object]:
    pro = create_tushare_client(token=token, http_url=http_url)
    index_df = filter_ths_index_targets(pro.ths_index())
    concept_codes = index_df["ts_code"].astype(str).tolist()
    concept_name_map = dict(zip(index_df["ts_code"], index_df["name"]))
    sync_config = {
        "job_name": job_name,
        "target_key": target_key,
        "start_date": start_date,
        "end_date": end_date,
        "concept_codes": concept_codes,
    }
    sleep_seconds = requests_per_min_to_sleep_seconds(requests_per_min)
    per_request_sleep_seconds = per_request_sleep_seconds if per_request_sleep_seconds is not None else sleep_seconds

    with psycopg.connect(db_dsn) as conn:
        state = load_sync_state(
            conn,
            job_name=str(sync_config["job_name"]),
            target_table="raw_ths_concept_daily",
            target_key=str(sync_config["target_key"]),
        )
        last_cursor = state["last_success_cursor"] if state else None

        def _fetch(concept_code: str, planned_start_date: str, planned_end_date: str) -> dict[str, pd.DataFrame]:
            return fetch_concept_daily_bundle(
                pro=pro,
                concept_code=concept_code,
                concept_name=concept_name_map[concept_code],
                start_date=planned_start_date,
                end_date=planned_end_date,
                per_request_sleep_seconds=per_request_sleep_seconds,
            )

        def _persist_frames(concept_code: str, frames: dict[str, pd.DataFrame]) -> None:
            persist_frames_to_db(conn, concept_code, frames)

        def _persist_state(payload: dict[str, object]) -> None:
            persist_sync_state_row(conn, payload)

        run_concept_daily_sync(
            sync_config=sync_config,
            last_cursor=last_cursor,
            fetch_concept_daily_bundle=_fetch,
            persist_frames=_persist_frames,
            persist_sync_state=_persist_state,
            end_date=end_date,
            overlap_days=overlap_days,
            sleep_seconds=0.0,
        )
    return sync_config


def sync_all_concept_members_to_db(
    start_date: str,
    end_date: str,
    token: str,
    job_name: str = DEFAULT_ALL_CONCEPT_MEMBERS_JOB_NAME,
    target_key: str = DEFAULT_ALL_CONCEPT_MEMBERS_TARGET_KEY,
    db_dsn: str = DEFAULT_DB_DSN,
    http_url: str = DEFAULT_TUSHARE_HTTP_URL,
    requests_per_min: int = DEFAULT_FEATURE_REQUESTS_PER_MIN,
    overlap_days: int = 7,
    per_request_sleep_seconds: float | None = None,
) -> dict[str, object]:
    del start_date, overlap_days
    pro = create_tushare_client(token=token, http_url=http_url)
    index_df = filter_ths_index_targets(pro.ths_index())
    concept_codes = index_df["ts_code"].astype(str).tolist()
    concept_name_map = dict(zip(index_df["ts_code"], index_df["name"]))
    sync_config = {
        "job_name": job_name,
        "target_key": target_key,
        "start_date": end_date,
        "end_date": end_date,
        "concept_codes": concept_codes,
    }
    sleep_seconds = requests_per_min_to_sleep_seconds(requests_per_min)
    per_request_sleep_seconds = per_request_sleep_seconds if per_request_sleep_seconds is not None else sleep_seconds
    mapping_asof_date = end_date

    with psycopg.connect(db_dsn) as conn:
        state = load_sync_state(
            conn,
            job_name=str(sync_config["job_name"]),
            target_table="raw_ths_member",
            target_key=str(sync_config["target_key"]),
        )
        last_cursor = state["last_success_cursor"] if state else None

        def _fetch(concept_code: str) -> dict[str, pd.DataFrame]:
            return fetch_concept_member_bundle(
                pro=pro,
                concept_code=concept_code,
                concept_name=concept_name_map[concept_code],
                mapping_asof_date=mapping_asof_date,
                per_request_sleep_seconds=per_request_sleep_seconds,
            )

        def _persist_frames(concept_code: str, frames: dict[str, pd.DataFrame]) -> None:
            persist_frames_to_db(conn, concept_code, frames)

        def _persist_state(payload: dict[str, object]) -> None:
            persist_sync_state_row(conn, payload)

        run_concept_member_sync(
            sync_config=sync_config,
            last_cursor=last_cursor,
            fetch_concept_member_bundle=_fetch,
            persist_frames=_persist_frames,
            persist_sync_state=_persist_state,
            end_date=end_date,
            sleep_seconds=0.0,
        )
    return sync_config


def sync_case_stock_bundle_to_db(
    workbook_path: str | Path,
    start_date: str,
    end_date: str,
    token: str,
    db_dsn: str = DEFAULT_DB_DSN,
    http_url: str = DEFAULT_TUSHARE_HTTP_URL,
    requests_per_min: int = 20,
    overlap_days: int = 7,
    exclude_names: list[str] | None = None,
    per_request_sleep_seconds: float | None = None,
) -> dict[str, object]:
    pro = create_tushare_client(token=token, http_url=http_url)
    stock_basic_df = pro.stock_basic(exchange="", list_status="L", fields="ts_code,name")
    stock_names = load_case_stock_names(workbook_path)
    sync_config = build_case_stock_sync_config(
        stock_names,
        stock_basic_df,
        start_date=start_date,
        exclude_names=exclude_names,
    )
    sleep_seconds = requests_per_min_to_sleep_seconds(requests_per_min)
    per_request_sleep_seconds = per_request_sleep_seconds if per_request_sleep_seconds is not None else sleep_seconds
    with psycopg.connect(db_dsn) as conn:
        state = load_sync_state(
            conn,
            job_name=str(sync_config["job_name"]),
            target_table="raw_stock_daily_qfq",
            target_key=str(sync_config["target_key"]),
        )
        last_cursor = state["last_success_cursor"] if state else None

        def _fetch(ts_code: str, planned_start_date: str, planned_end_date: str) -> dict[str, pd.DataFrame]:
            return fetch_case_stock_bundle(
                pro,
                ts_code,
                planned_start_date,
                planned_end_date,
                per_request_sleep_seconds=per_request_sleep_seconds,
            )

        def _persist_frames(ts_code: str, frames: dict[str, pd.DataFrame]) -> None:
            persist_frames_to_db(conn, ts_code, frames)

        def _persist_state(payload: dict[str, object]) -> None:
            persist_sync_state_row(conn, payload)

        run_case_stock_bundle_sync(
            sync_config=sync_config,
            last_cursor=last_cursor,
            fetch_stock_bundle=_fetch,
            persist_frames=_persist_frames,
            persist_sync_state=_persist_state,
            end_date=end_date,
            overlap_days=overlap_days,
            sleep_seconds=sleep_seconds,
        )
    return sync_config


def sync_stock_file_bundle_to_db(
    stock_file_path: str | Path,
    start_date: str,
    end_date: str,
    token: str,
    job_name: str,
    target_key: str,
    db_dsn: str = DEFAULT_DB_DSN,
    http_url: str = DEFAULT_TUSHARE_HTTP_URL,
    requests_per_min: int = 20,
    overlap_days: int = 7,
    code_column: str = "股票代码",
    per_request_sleep_seconds: float | None = None,
) -> dict[str, object]:
    pro = create_tushare_client(token=token, http_url=http_url)
    stock_codes = load_stock_codes_from_csv(stock_file_path, code_column=code_column)
    sync_config = build_stock_bundle_sync_config(
        stock_codes=stock_codes,
        start_date=start_date,
        job_name=job_name,
        target_key=target_key,
    )
    sleep_seconds = requests_per_min_to_sleep_seconds(requests_per_min)
    per_request_sleep_seconds = per_request_sleep_seconds if per_request_sleep_seconds is not None else sleep_seconds
    with psycopg.connect(db_dsn) as conn:
        state = load_sync_state(
            conn,
            job_name=str(sync_config["job_name"]),
            target_table="raw_stock_daily_qfq",
            target_key=str(sync_config["target_key"]),
        )
        last_cursor = state["last_success_cursor"] if state else None

        def _fetch(ts_code: str, planned_start_date: str, planned_end_date: str) -> dict[str, pd.DataFrame]:
            return fetch_case_stock_bundle(
                pro,
                ts_code,
                planned_start_date,
                planned_end_date,
                per_request_sleep_seconds=per_request_sleep_seconds,
            )

        def _persist_frames(ts_code: str, frames: dict[str, pd.DataFrame]) -> None:
            persist_frames_to_db(conn, ts_code, frames)

        def _persist_state(payload: dict[str, object]) -> None:
            persist_sync_state_row(conn, payload)

        run_case_stock_bundle_sync(
            sync_config=sync_config,
            last_cursor=last_cursor,
            fetch_stock_bundle=_fetch,
            persist_frames=_persist_frames,
            persist_sync_state=_persist_state,
            end_date=end_date,
            overlap_days=overlap_days,
            sleep_seconds=0.0,
        )
    return sync_config


def sync_all_stocks_by_trade_date_to_db(
    start_date: str,
    end_date: str,
    token: str,
    job_name: str = DEFAULT_ALL_STOCKS_JOB_NAME,
    target_key: str = DEFAULT_ALL_STOCKS_TARGET_KEY,
    db_dsn: str = DEFAULT_DB_DSN,
    http_url: str = DEFAULT_TUSHARE_HTTP_URL,
    requests_per_min: int = DEFAULT_FEATURE_REQUESTS_PER_MIN,
    overlap_days: int = 7,
    per_request_sleep_seconds: float | None = None,
    max_workers: int = DEFAULT_MAX_WORKERS,
) -> dict[str, object]:
    pro = create_tushare_client(token=token, http_url=http_url)
    trade_dates = list_open_trade_dates(pro, start_date=start_date, end_date=end_date)
    if not trade_dates:
        raise ValueError(f"未找到 {start_date} 到 {end_date} 的开市交易日")
    sleep_seconds = requests_per_min_to_sleep_seconds(requests_per_min)
    per_request_sleep_seconds = per_request_sleep_seconds if per_request_sleep_seconds is not None else sleep_seconds
    latest_adj_factor_map = load_latest_adj_factor_snapshot(
        pro,
        anchor_trade_date=trade_dates[-1],
        per_request_sleep_seconds=per_request_sleep_seconds,
    )
    sync_config = {
        "job_name": job_name,
        "target_key": target_key,
        "start_date": start_date,
        "end_date": end_date,
        "trade_dates": trade_dates,
    }
    with psycopg.connect(db_dsn) as conn:
        state = load_sync_state(
            conn,
            job_name=str(sync_config["job_name"]),
            target_table="raw_stock_daily_qfq",
            target_key=str(sync_config["target_key"]),
        )
        last_cursor = state["last_success_cursor"] if state else None

        def _fetch(trade_date: str) -> dict[str, pd.DataFrame]:
            return fetch_market_trade_date_bundle(
                pro,
                trade_date=trade_date,
                latest_adj_factor_map=latest_adj_factor_map,
                per_request_sleep_seconds=per_request_sleep_seconds,
                max_workers=max_workers,
            )

        def _persist_frames(trade_date: str, frames: dict[str, pd.DataFrame]) -> None:
            persist_frames_to_db(conn, trade_date, frames)

        def _persist_state(payload: dict[str, object]) -> None:
            persist_sync_state_row(conn, payload)

        run_market_trade_date_sync(
            sync_config=sync_config,
            last_cursor=last_cursor,
            fetch_trade_date_bundle=_fetch,
            persist_frames=_persist_frames,
            persist_sync_state=_persist_state,
            overlap_days=overlap_days,
        )
    return sync_config


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="同步 event_quant 的 Tushare 数据")
    subparsers = parser.add_subparsers(dest="command", required=True)

    case_parser = subparsers.add_parser("sync-case-stocks", help="同步 stock.xlsx 中案例股票包")
    case_parser.add_argument("--workbook", required=True, help="stock.xlsx 路径")
    case_parser.add_argument("--start-date", required=True, help="开始日期，格式 YYYYMMDD")
    case_parser.add_argument("--end-date", required=True, help="结束日期，格式 YYYYMMDD")
    case_parser.add_argument("--token", default=None, help="Tushare token")
    case_parser.add_argument("--db-dsn", default=None, help="PostgreSQL DSN")
    case_parser.add_argument("--http-url", default=None, help="Tushare 代理地址")
    case_parser.add_argument("--requests-per-min", type=int, default=20, help="每分钟请求上限")
    case_parser.add_argument("--overlap-days", type=int, default=7, help="断点续传回补天数")
    case_parser.add_argument("--exclude-name", action="append", default=[], help="排除指定标的名称，可重复传入")

    file_parser = subparsers.add_parser("sync-stock-file", help="同步股票清单文件中的股票包")
    file_parser.add_argument("--stock-file", required=True, help="股票清单 CSV 路径")
    file_parser.add_argument("--start-date", required=True, help="开始日期，格式 YYYYMMDD")
    file_parser.add_argument("--end-date", required=True, help="结束日期，格式 YYYYMMDD")
    file_parser.add_argument("--job-name", required=True, help="同步任务名")
    file_parser.add_argument("--target-key", required=True, help="同步目标键")
    file_parser.add_argument("--code-column", default="股票代码", help="股票代码列名")
    file_parser.add_argument("--token", default=None, help="Tushare token")
    file_parser.add_argument("--db-dsn", default=None, help="PostgreSQL DSN")
    file_parser.add_argument("--http-url", default=None, help="Tushare 代理地址")
    file_parser.add_argument("--requests-per-min", type=int, default=20, help="每分钟请求上限")
    file_parser.add_argument("--overlap-days", type=int, default=7, help="断点续传回补天数")

    concept_parser = subparsers.add_parser("sync-stock-file-concepts", help="同步股票清单文件对应的概念映射与概念指数")
    concept_parser.add_argument("--stock-file", required=True, help="股票清单 CSV 路径")
    concept_parser.add_argument("--start-date", required=True, help="开始日期，格式 YYYYMMDD")
    concept_parser.add_argument("--end-date", required=True, help="结束日期，格式 YYYYMMDD")
    concept_parser.add_argument("--job-name", required=True, help="同步任务名")
    concept_parser.add_argument("--target-key", required=True, help="同步目标键")
    concept_parser.add_argument("--code-column", default="股票代码", help="股票代码列名")
    concept_parser.add_argument("--token", default=None, help="Tushare token")
    concept_parser.add_argument("--db-dsn", default=None, help="PostgreSQL DSN")
    concept_parser.add_argument("--http-url", default=None, help="Tushare 代理地址")
    concept_parser.add_argument("--requests-per-min", type=int, default=20, help="每分钟请求上限")
    concept_parser.add_argument("--overlap-days", type=int, default=7, help="断点续传回补天数")

    all_stocks_parser = subparsers.add_parser("sync-all-stocks", help="按交易日同步全市场股票包")
    all_stocks_parser.add_argument("--start-date", required=True, help="开始日期，格式 YYYYMMDD")
    all_stocks_parser.add_argument("--end-date", required=True, help="结束日期，格式 YYYYMMDD")
    all_stocks_parser.add_argument("--job-name", default=DEFAULT_ALL_STOCKS_JOB_NAME, help="同步任务名")
    all_stocks_parser.add_argument("--target-key", default=DEFAULT_ALL_STOCKS_TARGET_KEY, help="同步目标键")
    all_stocks_parser.add_argument("--token", default=None, help="Tushare token")
    all_stocks_parser.add_argument("--db-dsn", default=None, help="PostgreSQL DSN")
    all_stocks_parser.add_argument("--http-url", default=None, help="Tushare 代理地址")
    all_stocks_parser.add_argument("--requests-per-min", type=int, default=DEFAULT_FEATURE_REQUESTS_PER_MIN, help="每分钟请求上限")
    all_stocks_parser.add_argument("--overlap-days", type=int, default=7, help="断点续传回补天数")
    all_stocks_parser.add_argument("--max-workers", type=int, default=DEFAULT_MAX_WORKERS, help="单交易日并发请求数")

    all_concepts_parser = subparsers.add_parser("sync-all-concepts", help="同步全量同花顺概念指数")
    all_concepts_parser.add_argument("--start-date", required=True, help="开始日期，格式 YYYYMMDD")
    all_concepts_parser.add_argument("--end-date", required=True, help="结束日期，格式 YYYYMMDD")
    all_concepts_parser.add_argument("--job-name", default=DEFAULT_ALL_CONCEPTS_JOB_NAME, help="同步任务名")
    all_concepts_parser.add_argument("--target-key", default=DEFAULT_ALL_CONCEPTS_TARGET_KEY, help="同步目标键")
    all_concepts_parser.add_argument("--token", default=None, help="Tushare token")
    all_concepts_parser.add_argument("--db-dsn", default=None, help="PostgreSQL DSN")
    all_concepts_parser.add_argument("--http-url", default=None, help="Tushare 代理地址")
    all_concepts_parser.add_argument("--requests-per-min", type=int, default=DEFAULT_FEATURE_REQUESTS_PER_MIN, help="每分钟请求上限")
    all_concepts_parser.add_argument("--overlap-days", type=int, default=7, help="断点续传回补天数")

    all_concept_members_parser = subparsers.add_parser("sync-all-concept-members", help="同步全量同花顺概念成员映射")
    all_concept_members_parser.add_argument("--start-date", required=True, help="开始日期，格式 YYYYMMDD")
    all_concept_members_parser.add_argument("--end-date", required=True, help="结束日期，格式 YYYYMMDD")
    all_concept_members_parser.add_argument("--job-name", default=DEFAULT_ALL_CONCEPT_MEMBERS_JOB_NAME, help="同步任务名")
    all_concept_members_parser.add_argument("--target-key", default=DEFAULT_ALL_CONCEPT_MEMBERS_TARGET_KEY, help="同步目标键")
    all_concept_members_parser.add_argument("--token", default=None, help="Tushare token")
    all_concept_members_parser.add_argument("--db-dsn", default=None, help="PostgreSQL DSN")
    all_concept_members_parser.add_argument("--http-url", default=None, help="Tushare 代理地址")
    all_concept_members_parser.add_argument("--requests-per-min", type=int, default=DEFAULT_FEATURE_REQUESTS_PER_MIN, help="每分钟请求上限")
    all_concept_members_parser.add_argument("--overlap-days", type=int, default=7, help="断点续传回补天数")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_argument_parser()
    args = parser.parse_args(argv)
    config = load_project_config(PROJECT_ROOT)

    token = args.token or config["tushare"]["token"]
    db_dsn = args.db_dsn or config["postgres"]["event_quant_dsn"]
    http_url = args.http_url or config["tushare"]["http_url"]

    if args.command == "sync-case-stocks":
        if not token:
            parser.error("缺少 --token 或项目配置 tushare.token")
        sync_config = sync_case_stock_bundle_to_db(
            workbook_path=args.workbook,
            start_date=args.start_date,
            end_date=args.end_date,
            token=token,
            db_dsn=db_dsn,
            http_url=http_url,
            requests_per_min=args.requests_per_min,
            overlap_days=args.overlap_days,
            exclude_names=args.exclude_name,
        )
        logger.info("案例股票包同步完成: %s", sync_config)
        return 0
    if args.command == "sync-stock-file":
        if not token:
            parser.error("缺少 --token 或项目配置 tushare.token")
        sync_config = sync_stock_file_bundle_to_db(
            stock_file_path=args.stock_file,
            start_date=args.start_date,
            end_date=args.end_date,
            token=token,
            job_name=args.job_name,
            target_key=args.target_key,
            code_column=args.code_column,
            db_dsn=db_dsn,
            http_url=http_url,
            requests_per_min=args.requests_per_min,
            overlap_days=args.overlap_days,
        )
        logger.info("股票清单同步完成: %s", sync_config)
        return 0
    if args.command == "sync-stock-file-concepts":
        if not token:
            parser.error("缺少 --token 或项目配置 tushare.token")
        sync_config = sync_stock_file_concepts_to_db(
            stock_file_path=args.stock_file,
            start_date=args.start_date,
            end_date=args.end_date,
            token=token,
            job_name=args.job_name,
            target_key=args.target_key,
            code_column=args.code_column,
            db_dsn=db_dsn,
            http_url=http_url,
            requests_per_min=args.requests_per_min,
            overlap_days=args.overlap_days,
        )
        logger.info("股票清单概念同步完成: %s", sync_config)
        return 0
    if args.command == "sync-all-stocks":
        if not token:
            parser.error("缺少 --token 或项目配置 tushare.token")
        sync_config = sync_all_stocks_by_trade_date_to_db(
            start_date=args.start_date,
            end_date=args.end_date,
            token=token,
            job_name=args.job_name,
            target_key=args.target_key,
            db_dsn=db_dsn,
            http_url=http_url,
            requests_per_min=args.requests_per_min,
            overlap_days=args.overlap_days,
            max_workers=args.max_workers,
        )
        logger.info("全市场股票包同步完成: %s", sync_config)
        return 0
    if args.command == "sync-all-concepts":
        if not token:
            parser.error("缺少 --token 或项目配置 tushare.token")
        sync_config = sync_all_concepts_to_db(
            start_date=args.start_date,
            end_date=args.end_date,
            token=token,
            job_name=args.job_name,
            target_key=args.target_key,
            db_dsn=db_dsn,
            http_url=http_url,
            requests_per_min=args.requests_per_min,
            overlap_days=args.overlap_days,
        )
        logger.info("全量概念指数同步完成: %s", sync_config)
        return 0
    if args.command == "sync-all-concept-members":
        if not token:
            parser.error("缺少 --token 或项目配置 tushare.token")
        sync_config = sync_all_concept_members_to_db(
            start_date=args.start_date,
            end_date=args.end_date,
            token=token,
            job_name=args.job_name,
            target_key=args.target_key,
            db_dsn=db_dsn,
            http_url=http_url,
            requests_per_min=args.requests_per_min,
            overlap_days=args.overlap_days,
        )
        logger.info("全量概念成员映射同步完成: %s", sync_config)
        return 0
    parser.error(f"不支持的命令: {args.command}")
    return 1

def filter_ths_index_targets(index_df: pd.DataFrame) -> pd.DataFrame:
    allowed = index_df[index_df["type"].isin(["N", "I"])].copy()
    allowed = allowed[allowed["ts_code"].astype(str).str.startswith(THEME_CONCEPT_PREFIXES)].copy()
    return allowed.reset_index(drop=True)


def build_theme_concept_matches(
    index_df: pd.DataFrame,
    stock_codes: list[str],
    fetch_members: Callable[[str], pd.DataFrame],
    mapping_asof_date: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    stock_set = set(stock_codes)
    member_frames: list[pd.DataFrame] = []
    matched_concepts: list[dict[str, str]] = []
    filtered = filter_ths_index_targets(index_df)
    for row in filtered.itertuples(index=False):
        member_df = fetch_members(row.ts_code)
        if member_df is None or member_df.empty:
            continue
        hits = member_df[member_df["con_code"].isin(stock_set)].copy()
        if hits.empty:
            continue
        member_frames.append(normalize_ths_member(hits, mapping_asof_date))
        matched_concepts.append({"concept_code": row.ts_code, "concept_name": row.name})

    if not member_frames:
        empty_member = pd.DataFrame(columns=["ts_code", "con_code", "con_name", "mapping_asof_date"])
        empty_concepts = pd.DataFrame(columns=["concept_code", "concept_name"])
        return empty_member, empty_concepts

    member_frame = pd.concat(member_frames, ignore_index=True).drop_duplicates().reset_index(drop=True)
    concept_frame = pd.DataFrame(matched_concepts).drop_duplicates().reset_index(drop=True)
    return member_frame, concept_frame


def normalize_stock_daily_qfq(df: pd.DataFrame) -> pd.DataFrame:
    result = df.rename(
        columns={
            "open": "open_qfq",
            "high": "high_qfq",
            "low": "low_qfq",
            "close": "close_qfq",
        }
    )[
        [
            "ts_code",
            "trade_date",
            "open_qfq",
            "high_qfq",
            "low_qfq",
            "close_qfq",
            "pct_chg",
            "vol",
            "amount",
        ]
    ].copy()
    return result


def resolve_akshare_stock_params(ts_code: str) -> tuple[str, str]:
    symbol, market_suffix = ts_code.split(".")
    market_map = {"SH": "sh", "SZ": "sz", "BJ": "bj"}
    return symbol, market_map.get(market_suffix.upper(), market_suffix.lower())


def _empty_stock_daily_qfq_frame() -> pd.DataFrame:
    return pd.DataFrame(
        columns=["ts_code", "trade_date", "open_qfq", "high_qfq", "low_qfq", "close_qfq", "pct_chg", "vol", "amount"]
    )


def _empty_daily_basic_frame() -> pd.DataFrame:
    return pd.DataFrame(
        columns=["ts_code", "trade_date", "turnover_rate", "turnover_rate_f", "volume_ratio", "pe", "pb", "total_mv", "circ_mv"]
    )


def _empty_moneyflow_frame() -> pd.DataFrame:
    return pd.DataFrame(
        columns=["ts_code", "trade_date", "buy_lg_amount", "sell_lg_amount", "buy_elg_amount", "sell_elg_amount", "net_mf_amount"]
    )


def _empty_limit_list_frame() -> pd.DataFrame:
    return pd.DataFrame(
        columns=["ts_code", "trade_date", "fd_amount", "first_time", "last_time", "open_times", "limit_status"]
    )


def build_akshare_qfq_daily_frame(hist_df: pd.DataFrame, ts_code: str) -> pd.DataFrame:
    if hist_df is None or hist_df.empty:
        return _empty_stock_daily_qfq_frame()
    result = pd.DataFrame(
        {
            "ts_code": ts_code,
            "trade_date": pd.to_datetime(hist_df["日期"]).dt.strftime("%Y-%m-%d"),
            "open": pd.to_numeric(hist_df["开盘"], errors="coerce"),
            "high": pd.to_numeric(hist_df["最高"], errors="coerce"),
            "low": pd.to_numeric(hist_df["最低"], errors="coerce"),
            "close": pd.to_numeric(hist_df["收盘"], errors="coerce"),
            "pct_chg": pd.to_numeric(hist_df["涨跌幅"], errors="coerce"),
            "vol": pd.to_numeric(hist_df["成交量"], errors="coerce"),
            "amount": pd.to_numeric(hist_df["成交额"], errors="coerce"),
        }
    ).sort_values("trade_date")
    return normalize_stock_daily_qfq(result)


def build_akshare_daily_basic_frame(hist_df: pd.DataFrame, ts_code: str) -> pd.DataFrame:
    if hist_df is None or hist_df.empty:
        return _empty_daily_basic_frame()
    result = pd.DataFrame(
        {
            "ts_code": ts_code,
            "trade_date": pd.to_datetime(hist_df["日期"]).dt.strftime("%Y-%m-%d"),
            "turnover_rate": pd.to_numeric(hist_df["换手率"], errors="coerce"),
            "turnover_rate_f": np.nan,
            "volume_ratio": np.nan,
            "pe": np.nan,
            "pb": np.nan,
            "total_mv": np.nan,
            "circ_mv": np.nan,
        }
    ).sort_values("trade_date")
    return normalize_daily_basic(result)


def build_akshare_moneyflow_frame(flow_df: pd.DataFrame, ts_code: str, start_date: str, end_date: str) -> pd.DataFrame:
    if flow_df is None or flow_df.empty:
        return _empty_moneyflow_frame()
    start_iso = datetime.strptime(start_date, "%Y%m%d").strftime("%Y-%m-%d")
    end_iso = datetime.strptime(end_date, "%Y%m%d").strftime("%Y-%m-%d")
    result = pd.DataFrame(
        {
            "ts_code": ts_code,
            "trade_date": pd.to_datetime(flow_df["日期"]).dt.strftime("%Y-%m-%d"),
            "buy_lg_amount": np.nan,
            "sell_lg_amount": np.nan,
            "buy_elg_amount": np.nan,
            "sell_elg_amount": np.nan,
            "net_mf_amount": pd.to_numeric(flow_df["主力净流入-净额"], errors="coerce"),
        }
    )
    result = result[(result["trade_date"] >= start_iso) & (result["trade_date"] <= end_iso)].sort_values("trade_date")
    if result.empty:
        return _empty_moneyflow_frame()
    return normalize_moneyflow(result)


def build_akshare_limit_list_frame(
    hist_df: pd.DataFrame,
    ts_code: str,
    ak_client,
) -> pd.DataFrame:
    if hist_df is None or hist_df.empty:
        return _empty_limit_list_frame()
    symbol, _ = resolve_akshare_stock_params(ts_code)
    trade_df = pd.DataFrame(
        {
            "trade_date": pd.to_datetime(hist_df["日期"]).dt.strftime("%Y%m%d"),
            "pct_chg": pd.to_numeric(hist_df["涨跌幅"], errors="coerce"),
        }
    ).dropna()
    rows: list[dict[str, object]] = []
    positive_days = sorted(trade_df.loc[trade_df["pct_chg"] >= 4.5, "trade_date"].unique())
    for trade_date in positive_days:
        try:
            day_df = ak_client.stock_zt_pool_em(date=trade_date)
        except Exception as exc:
            logger.warning("Akshare 涨停池拉取失败，跳过该日: ts_code=%s trade_date=%s error=%s", ts_code, trade_date, exc)
            continue
        if day_df is None or day_df.empty:
            continue
        hit = day_df.loc[day_df["代码"].astype(str) == symbol]
        if hit.empty:
            continue
        row = hit.iloc[0]
        rows.append(
            {
                "ts_code": ts_code,
                "trade_date": datetime.strptime(trade_date, "%Y%m%d").strftime("%Y-%m-%d"),
                "fd_amount": row.get("封板资金"),
                "first_time": row.get("首次封板时间"),
                "last_time": row.get("最后封板时间"),
                "open_times": row.get("炸板次数", 0),
                "limit_status": "U",
            }
        )
    if not rows:
        return _empty_limit_list_frame()
    return normalize_limit_list_d(pd.DataFrame(rows).sort_values("trade_date"))


def fetch_case_stock_bundle_from_akshare(
    ts_code: str,
    start_date: str,
    end_date: str,
    ak_client=None,
) -> dict[str, pd.DataFrame]:
    ak_client = ak_client or ak
    if ak_client is None:
        raise RuntimeError("Akshare 不可用")
    symbol, market = resolve_akshare_stock_params(ts_code)
    logger.info(
        "开始抓取股票包[Akshare]: ts_code=%s symbol=%s market=%s start_date=%s end_date=%s",
        ts_code,
        symbol,
        market,
        start_date,
        end_date,
    )

    def _load_bundle() -> tuple[pd.DataFrame, pd.DataFrame]:
        hist = ak_client.stock_zh_a_hist(
            symbol=symbol,
            period="daily",
            start_date=start_date,
            end_date=end_date,
            adjust="qfq",
        )
        fund = ak_client.stock_individual_fund_flow(stock=symbol, market=market)
        return hist, fund

    try:
        with requests_sessions_without_proxy():
            hist_df, moneyflow_df = _load_bundle()
    except Exception:
        logger.warning(
            "Akshare 无代理请求失败，回退到默认网络环境: ts_code=%s symbol=%s",
            ts_code,
            symbol,
        )
        hist_df, moneyflow_df = _load_bundle()
    return {
        "raw_stock_daily_qfq": build_akshare_qfq_daily_frame(hist_df, ts_code),
        "raw_daily_basic": build_akshare_daily_basic_frame(hist_df, ts_code),
        "raw_moneyflow": build_akshare_moneyflow_frame(moneyflow_df, ts_code, start_date, end_date),
        "raw_limit_list_d": build_akshare_limit_list_frame(hist_df, ts_code, ak_client),
    }


def _empty_ana_stock_concept_map_frame() -> pd.DataFrame:
    return pd.DataFrame(
        columns=["ts_code", "concept_code", "concept_name", "mapping_asof_date", "map_source", "updated_at"]
    )


def _empty_ana_concept_day_frame() -> pd.DataFrame:
    return pd.DataFrame(columns=["concept_code", "trade_date", "concept_name", "close", "pct_change", "vol", "turnover_rate"])


def normalize_em_stock_concept_map(
    *,
    ts_code: str,
    concept_code: str,
    concept_name: str,
    mapping_asof_date: str,
    updated_at: str | None = None,
) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "ts_code": ts_code,
                "concept_code": concept_code,
                "concept_name": concept_name,
                "mapping_asof_date": pd.to_datetime(mapping_asof_date).strftime("%Y-%m-%d"),
                "map_source": "akshare_em_concept",
                "updated_at": updated_at or datetime.now().isoformat(),
            }
        ]
    )


def normalize_em_concept_day(df: pd.DataFrame, concept_code: str, concept_name: str) -> pd.DataFrame:
    if df is None or df.empty:
        return _empty_ana_concept_day_frame()
    result = df.rename(
        columns={
            "日期": "trade_date",
            "收盘": "close",
            "涨跌幅": "pct_change",
            "成交量": "vol",
            "换手率": "turnover_rate",
        }
    )[["trade_date", "close", "pct_change", "vol", "turnover_rate"]].copy()
    result.insert(0, "concept_code", concept_code)
    result.insert(2, "concept_name", concept_name)
    result["trade_date"] = pd.to_datetime(result["trade_date"]).dt.strftime("%Y-%m-%d")
    return result


def fetch_case_stock_concept_bundle_from_akshare(
    ts_code: str,
    start_date: str,
    end_date: str,
    concept_name_fetcher: Callable[[], pd.DataFrame] | None = None,
    concept_cons_fetcher: Callable[..., pd.DataFrame] | None = None,
    concept_hist_fetcher: Callable[..., pd.DataFrame] | None = None,
    ak_client=None,
    per_request_sleep_seconds: float = 0.0,
    sleeper: Callable[[float], None] | None = None,
) -> dict[str, pd.DataFrame]:
    ak_client = ak_client or ak
    if ak_client is None:
        raise RuntimeError("Akshare 不可用")

    concept_name_fetcher = concept_name_fetcher or ak_client.stock_board_concept_name_em
    concept_cons_fetcher = concept_cons_fetcher or ak_client.stock_board_concept_cons_em
    concept_hist_fetcher = concept_hist_fetcher or ak_client.stock_board_concept_hist_em
    sleeper = sleeper or time.sleep

    symbol, _ = resolve_akshare_stock_params(ts_code)
    logger.info(
        "开始抓取概念包[Akshare]: ts_code=%s symbol=%s start_date=%s end_date=%s",
        ts_code,
        symbol,
        start_date,
        end_date,
    )

    def _load_concept_names() -> pd.DataFrame:
        return concept_name_fetcher()

    try:
        with requests_sessions_without_proxy():
            concept_name_df = call_with_rate_limit_retry(_load_concept_names)
    except Exception:
        logger.warning(
            "Akshare 概念列表无代理请求失败，回退到默认网络环境: ts_code=%s symbol=%s",
            ts_code,
            symbol,
        )
        concept_name_df = call_with_rate_limit_retry(_load_concept_names)

    if concept_name_df is None or concept_name_df.empty:
        return {
            "ana_stock_concept_map": _empty_ana_stock_concept_map_frame(),
            "ana_concept_day": _empty_ana_concept_day_frame(),
        }

    concept_name_df = concept_name_df.dropna(subset=["板块代码", "板块名称"]).copy()
    mapping_frames: list[pd.DataFrame] = []
    concept_day_frames: list[pd.DataFrame] = []

    for row in concept_name_df[["板块代码", "板块名称"]].itertuples(index=False):
        concept_code = str(row.板块代码)
        concept_name = str(row.板块名称)

        def _load_cons() -> pd.DataFrame:
            return concept_cons_fetcher(symbol=concept_code)

        cons_df = call_with_rate_limit_retry(_load_cons)
        if per_request_sleep_seconds > 0:
            sleeper(per_request_sleep_seconds)
        if cons_df is None or cons_df.empty or "代码" not in cons_df.columns:
            continue
        codes = cons_df["代码"].astype(str).str.extract(r"(\d+)")[0].fillna("").str.zfill(6)
        if symbol not in set(codes.tolist()):
            continue

        mapping_frames.append(
            normalize_em_stock_concept_map(
                ts_code=ts_code,
                concept_code=concept_code,
                concept_name=concept_name,
                mapping_asof_date=end_date,
            )
        )

        def _load_hist() -> pd.DataFrame:
            return concept_hist_fetcher(
                symbol=concept_name,
                period="daily",
                start_date=start_date,
                end_date=end_date,
                adjust="",
            )

        hist_df = call_with_rate_limit_retry(_load_hist)
        if per_request_sleep_seconds > 0:
            sleeper(per_request_sleep_seconds)
        concept_day_frames.append(normalize_em_concept_day(hist_df, concept_code, concept_name))

    if not mapping_frames:
        return {
            "ana_stock_concept_map": _empty_ana_stock_concept_map_frame(),
            "ana_concept_day": _empty_ana_concept_day_frame(),
        }

    return {
        "ana_stock_concept_map": pd.concat(mapping_frames, ignore_index=True)
        .drop_duplicates(subset=["ts_code", "concept_code"])
        .reset_index(drop=True),
        "ana_concept_day": pd.concat(concept_day_frames, ignore_index=True)
        .drop_duplicates(subset=["concept_code", "trade_date"])
        .reset_index(drop=True)
        if concept_day_frames
        else _empty_ana_concept_day_frame(),
    }


def build_qfq_daily_frame(daily_df: pd.DataFrame, adj_factor_df: pd.DataFrame) -> pd.DataFrame:
    if daily_df is None or daily_df.empty:
        return pd.DataFrame(
            columns=[
                "ts_code",
                "trade_date",
                "open_qfq",
                "high_qfq",
                "low_qfq",
                "close_qfq",
                "pct_chg",
                "vol",
                "amount",
            ]
        )

    required_columns = {"ts_code", "trade_date", "adj_factor"}
    if adj_factor_df is None or adj_factor_df.empty or not required_columns.issubset(set(adj_factor_df.columns)):
        logger.warning("adj_factor 缺失或列不完整，回退为原始日线价格", extra={"rows": len(daily_df)})
        return normalize_stock_daily_qfq(daily_df)

    merged = daily_df.merge(
        adj_factor_df[["ts_code", "trade_date", "adj_factor"]],
        on=["ts_code", "trade_date"],
        how="left",
    )
    merged = merged.sort_values("trade_date", ascending=False).reset_index(drop=True)
    latest_adj_factor = merged.iloc[0]["adj_factor"]
    factor_scale = merged["adj_factor"] / latest_adj_factor

    merged["open_qfq"] = merged["open"] * factor_scale
    merged["high_qfq"] = merged["high"] * factor_scale
    merged["low_qfq"] = merged["low"] * factor_scale
    merged["close_qfq"] = merged["close"] * factor_scale

    return merged[
        [
            "ts_code",
            "trade_date",
            "open_qfq",
            "high_qfq",
            "low_qfq",
            "close_qfq",
            "pct_chg",
            "vol",
            "amount",
        ]
    ].copy()


def normalize_ths_member(df: pd.DataFrame, mapping_asof_date: str) -> pd.DataFrame:
    result = df[["ts_code", "con_code", "con_name"]].copy()
    result["mapping_asof_date"] = pd.to_datetime(mapping_asof_date).strftime("%Y-%m-%d")
    return result


def normalize_index_daily(df: pd.DataFrame) -> pd.DataFrame:
    return df[
        [
            "ts_code",
            "trade_date",
            "open",
            "high",
            "low",
            "close",
            "pct_chg",
            "vol",
            "amount",
        ]
    ].copy()


def normalize_ths_concept_daily(df: pd.DataFrame, concept_name: str) -> pd.DataFrame:
    result = df[
        [
            "ts_code",
            "trade_date",
            "open",
            "high",
            "low",
            "close",
            "pct_change",
            "vol",
            "turnover_rate",
        ]
    ].copy()
    result.insert(2, "concept_name", concept_name)
    return result


def normalize_ana_stock_concept_map(
    member_frame: pd.DataFrame,
    concept_meta: pd.DataFrame,
    updated_at: str | None = None,
) -> pd.DataFrame:
    if member_frame is None or member_frame.empty:
        return pd.DataFrame(
            columns=["ts_code", "concept_code", "concept_name", "mapping_asof_date", "map_source", "updated_at"]
        )
    concept_name_map = dict(zip(concept_meta["concept_code"], concept_meta["concept_name"]))
    result = pd.DataFrame(
        {
            "ts_code": member_frame["con_code"],
            "concept_code": member_frame["ts_code"],
            "concept_name": member_frame["ts_code"].map(concept_name_map),
            "mapping_asof_date": member_frame["mapping_asof_date"],
            "map_source": "ths_member",
            "updated_at": updated_at or datetime.now().isoformat(),
        }
    )
    return result.drop_duplicates().reset_index(drop=True)


def normalize_ana_concept_day(raw_df: pd.DataFrame) -> pd.DataFrame:
    if raw_df is None or raw_df.empty:
        return pd.DataFrame(columns=["concept_code", "trade_date", "concept_name", "close", "pct_change", "vol", "turnover_rate"])
    return raw_df.rename(columns={"ts_code": "concept_code"})[
        ["concept_code", "trade_date", "concept_name", "close", "pct_change", "vol", "turnover_rate"]
    ].copy()


def normalize_daily_basic(df: pd.DataFrame) -> pd.DataFrame:
    return df[
        [
            "ts_code",
            "trade_date",
            "turnover_rate",
            "turnover_rate_f",
            "volume_ratio",
            "pe",
            "pb",
            "total_mv",
            "circ_mv",
        ]
    ].copy()


def normalize_moneyflow(df: pd.DataFrame) -> pd.DataFrame:
    return df[
        [
            "ts_code",
            "trade_date",
            "buy_lg_amount",
            "sell_lg_amount",
            "buy_elg_amount",
            "sell_elg_amount",
            "net_mf_amount",
        ]
    ].copy()


def normalize_limit_list_d(df: pd.DataFrame) -> pd.DataFrame:
    return df.rename(columns={"limit": "limit_status"})[
        [
            "ts_code",
            "trade_date",
            "fd_amount",
            "first_time",
            "last_time",
            "open_times",
            "limit_status",
        ]
    ].copy()


if __name__ == "__main__":
    raise SystemExit(main())
