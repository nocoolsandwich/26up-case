from __future__ import annotations

import argparse
import json
import os
import time
from pathlib import Path

import akshare as ak
import pandas as pd
import requests
from openpyxl import load_workbook

TODAY = "20260306"
CACHE_DIR = Path("data/stock_cache")
CODE_CACHE_PATH = CACHE_DIR / "a_share_code_name.csv"
REPORT_PATH = Path("data/stock_case_periods.csv")


def _without_proxy_env() -> None:
    for key in ["HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY", "http_proxy", "https_proxy", "all_proxy"]:
        os.environ.pop(key, None)
    os.environ["NO_PROXY"] = "*"
    os.environ["no_proxy"] = "*"
    requests.sessions.Session.trust_env = False


def ensure_cache_dir(cache_dir: Path = CACHE_DIR) -> None:
    cache_dir.mkdir(parents=True, exist_ok=True)


def load_code_map(cache_path: Path = CODE_CACHE_PATH, refresh: bool = False) -> dict[str, str]:
    ensure_cache_dir(cache_path.parent)
    if cache_path.exists() and not refresh:
        df = pd.read_csv(cache_path, dtype=str)
    else:
        _without_proxy_env()
        df = ak.stock_info_a_code_name()
        df.to_csv(cache_path, index=False)
    name_col = "name" if "name" in df.columns else df.columns[1]
    code_col = "code" if "code" in df.columns else df.columns[0]
    return {str(row[name_col]).strip(): str(row[code_col]).strip() for _, row in df.iterrows()}


def _hist_cache_path(symbol: str, start_date: str, end_date: str, cache_dir: Path = CACHE_DIR) -> Path:
    return cache_dir / f"{symbol}_{start_date}_{end_date}_qfq.csv"


def _to_market_symbol(symbol: str) -> str:
    return f"sh{symbol}" if symbol.startswith("6") else f"sz{symbol}"


def load_stock_hist(symbol: str, start_date: str, end_date: str, cache_dir: Path = CACHE_DIR, refresh: bool = False) -> pd.DataFrame:
    ensure_cache_dir(cache_dir)
    cache_path = _hist_cache_path(symbol, start_date, end_date, cache_dir)
    if cache_path.exists() and not refresh:
        return pd.read_csv(cache_path)

    _without_proxy_env()
    last_err: Exception | None = None
    for _ in range(3):
        try:
            df = ak.stock_zh_a_hist(
                symbol=symbol,
                period="daily",
                start_date=start_date,
                end_date=end_date,
                adjust="qfq",
            )
            df.to_csv(cache_path, index=False)
            return df
        except Exception as exc:
            last_err = exc
            time.sleep(0.6)
    market_symbol = _to_market_symbol(symbol)
    for fetcher in (
        lambda: ak.stock_zh_a_hist_tx(
            symbol=market_symbol,
            start_date=start_date,
            end_date=end_date,
            adjust="qfq",
            timeout=10,
        ),
        lambda: ak.stock_zh_a_daily(
            symbol=market_symbol,
            start_date=start_date,
            end_date=end_date,
            adjust="qfq",
        ),
    ):
        for _ in range(2):
            try:
                df = fetcher()
                df.to_csv(cache_path, index=False)
                return df
            except Exception as exc:
                last_err = exc
                time.sleep(0.6)
    if last_err is not None:
        raise last_err
    raise RuntimeError(f"failed to fetch hist for {symbol}")


def resolve_case_period(hist_df: pd.DataFrame, has_existing_end: bool) -> dict[str, object]:
    if hist_df is None or hist_df.empty:
        raise ValueError("hist_df is empty")

    hist = hist_df.copy()
    hist["日期"] = pd.to_datetime(hist["日期"])
    hist = hist.sort_values("日期").reset_index(drop=True)
    low_row = hist.loc[hist["最低"].astype(float).idxmin()]
    if has_existing_end:
        end_row = hist.iloc[-1]
        mode = "existing_end_low_to_high"
    else:
        end_row = hist.loc[hist["最高"].astype(float).idxmax()]
        mode = "filled_end_low_to_high"

    start_price = float(low_row["最低"])
    end_price = float(end_row["最高"])
    max_profit_pct = (end_price / start_price - 1.0) * 100.0 if start_price else None
    return {
        "start_date": hist.iloc[0]["日期"].strftime("%Y-%m-%d"),
        "buy_date": pd.Timestamp(low_row["日期"]).strftime("%Y-%m-%d"),
        "end_date": pd.Timestamp(end_row["日期"]).strftime("%Y-%m-%d"),
        "start_price": start_price,
        "end_price": end_price,
        "max_profit_pct": max_profit_pct,
        "mode": mode,
    }


def build_case_period_report(
    input_path: Path,
    refresh: bool = False,
    today: str = TODAY,
    excluded_names: set[str] | None = None,
) -> pd.DataFrame:
    wb = pd.read_excel(input_path, sheet_name="案例库")
    code_map = load_code_map(refresh=refresh)
    excluded_names = excluded_names or set()
    rows: list[dict[str, object]] = []
    for _, row in wb.iterrows():
        case_id = row["case_id"]
        name = str(row["标的名称"]).strip()
        if name in excluded_names:
            continue
        start = pd.to_datetime(row["开始日期"]).strftime("%Y%m%d")
        end_value = row["结束日期"]
        has_existing_end = pd.notna(end_value)
        end = pd.to_datetime(end_value).strftime("%Y%m%d") if has_existing_end else today
        code = code_map.get(name, "")
        result: dict[str, object] = {
            "case_id": case_id,
            "stock_name": name,
            "stock_code": code,
            "input_start_date": pd.to_datetime(row["开始日期"]).strftime("%Y-%m-%d"),
            "input_end_date": pd.to_datetime(end_value).strftime("%Y-%m-%d") if has_existing_end else "",
        }
        if not code:
            result["error"] = "stock_code_not_found"
            rows.append(result)
            continue
        try:
            hist = load_stock_hist(code, start, end, refresh=refresh)
            period = resolve_case_period(hist, has_existing_end=has_existing_end)
            result.update(period)
        except Exception as exc:
            result["error"] = f"{type(exc).__name__}: {exc}"
        rows.append(result)
    return pd.DataFrame(rows)


def apply_report_to_workbook(
    workbook_path: Path,
    report_df: pd.DataFrame,
    sheet_name: str = "案例库",
    excluded_names: set[str] | None = None,
) -> int:
    excluded_names = excluded_names or set()
    usable = report_df.copy()
    if "error" in usable.columns:
        usable = usable[(usable["error"].fillna("") == "")]
    usable = usable[~usable["stock_name"].isin(excluded_names)]
    by_case_id = {str(row["case_id"]): row for _, row in usable.iterrows()}

    wb = load_workbook(workbook_path)
    ws = wb[sheet_name]
    headers = {str(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}
    case_id_col = headers["case_id"]
    end_date_col = headers["结束日期"]
    profit_col = headers["收益率(%)"]

    updated_rows = 0
    for row_idx in range(2, ws.max_row + 1):
        case_id = str(ws.cell(row_idx, case_id_col).value)
        stock_name = str(ws.cell(row_idx, headers["标的名称"]).value).strip()
        if stock_name in excluded_names:
            continue
        report_row = by_case_id.get(case_id)
        if report_row is None:
            continue
        end_date = report_row.get("end_date", "")
        max_profit_pct = report_row.get("max_profit_pct", "")
        ws.cell(row=row_idx, column=end_date_col).value = pd.to_datetime(end_date).to_pydatetime() if str(end_date).strip() else None
        ws.cell(row=row_idx, column=profit_col).value = float(max_profit_pct) if str(max_profit_pct).strip() else None
        updated_rows += 1

    wb.save(workbook_path)
    return updated_rows


def main() -> None:
    parser = argparse.ArgumentParser(description="Calculate case end dates and period returns with local cache")
    parser.add_argument("--input", default="stock.xlsx", help="Input xlsx path")
    parser.add_argument("--output", default=str(REPORT_PATH), help="Output csv path")
    parser.add_argument("--refresh", action="store_true", help="Refresh code map and K-line cache")
    parser.add_argument("--exclude-name", action="append", default=[], help="Exclude stock names from output")
    parser.add_argument("--writeback", action="store_true", help="Write end date and max profit back to workbook")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    report_df = build_case_period_report(
        input_path=input_path,
        refresh=args.refresh,
        excluded_names=set(args.exclude_name),
    )
    report_df.to_csv(output_path, index=False)
    updated_rows = 0
    if args.writeback:
        updated_rows = apply_report_to_workbook(
            workbook_path=input_path,
            report_df=report_df,
            excluded_names=set(args.exclude_name),
        )
    print(f"rows={len(report_df)}")
    print(f"output={output_path}")
    if args.writeback:
        print(f"updated_rows={updated_rows}")
    print(json.dumps(report_df.fillna('').to_dict(orient='records'), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
