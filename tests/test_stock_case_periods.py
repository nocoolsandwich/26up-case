import os
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from scripts.calc_stock_case_periods import (
    _without_proxy_env,
    apply_report_to_workbook,
    build_case_period_report,
    resolve_case_period,
)


class ResolveCasePeriodTest(unittest.TestCase):
    def test_without_proxy_env_clears_proxy_vars(self):
        os.environ["HTTP_PROXY"] = "http://127.0.0.1:7890"
        os.environ["https_proxy"] = "http://127.0.0.1:7890"

        _without_proxy_env()

        self.assertNotIn("HTTP_PROXY", os.environ)
        self.assertNotIn("https_proxy", os.environ)

    def test_keeps_existing_end_date_and_uses_range_low_to_range_high(self):
        hist = pd.DataFrame(
            [
                {"日期": "2026-02-26", "收盘": 10.0, "最高": 10.3, "最低": 9.8},
                {"日期": "2026-02-27", "收盘": 11.0, "最高": 11.5, "最低": 9.5},
                {"日期": "2026-03-02", "收盘": 12.0, "最高": 12.2, "最低": 11.7},
            ]
        )

        result = resolve_case_period(hist, has_existing_end=True)

        self.assertEqual(result["end_date"], "2026-03-02")
        self.assertEqual(result["buy_date"], "2026-02-27")
        self.assertAlmostEqual(result["start_price"], 9.5)
        self.assertAlmostEqual(result["end_price"], 12.2)
        self.assertAlmostEqual(result["max_profit_pct"], (12.2 / 9.5 - 1) * 100)
        self.assertEqual(result["mode"], "existing_end_low_to_high")

    def test_fills_missing_end_date_with_max_high_date(self):
        hist = pd.DataFrame(
            [
                {"日期": "2026-02-26", "收盘": 10.0, "最高": 10.3, "最低": 9.8},
                {"日期": "2026-02-27", "收盘": 11.0, "最高": 13.5, "最低": 10.5},
                {"日期": "2026-03-02", "收盘": 12.0, "最高": 12.2, "最低": 11.7},
            ]
        )

        result = resolve_case_period(hist, has_existing_end=False)

        self.assertEqual(result["end_date"], "2026-02-27")
        self.assertEqual(result["buy_date"], "2026-02-26")
        self.assertAlmostEqual(result["start_price"], 9.8)
        self.assertAlmostEqual(result["end_price"], 13.5)
        self.assertAlmostEqual(result["max_profit_pct"], (13.5 / 9.8 - 1) * 100)
        self.assertEqual(result["mode"], "filled_end_low_to_high")

    def test_build_case_period_report_keeps_running_when_one_stock_fetch_fails(self):
        with TemporaryDirectory() as tmpdir:
            input_path = Path(tmpdir) / "stock.xlsx"
            df = pd.DataFrame(
                [
                    {"case_id": "案例001", "标的名称": "五洲新春", "开始日期": "2025-11-05", "结束日期": "2026-01-22"},
                    {"case_id": "案例002", "标的名称": "信维通信", "开始日期": "2025-12-24", "结束日期": pd.NaT},
                ]
            )
            with pd.ExcelWriter(input_path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="案例库", index=False)

            ok_hist = pd.DataFrame(
                [
                    {"日期": "2025-11-05", "收盘": 10.0, "最高": 10.5, "最低": 9.9},
                    {"日期": "2026-01-22", "收盘": 12.0, "最高": 12.2, "最低": 11.8},
                ]
            )

            with patch("scripts.calc_stock_case_periods.load_code_map", return_value={"五洲新春": "603667", "信维通信": "300136"}):
                with patch("scripts.calc_stock_case_periods.load_stock_hist", side_effect=[ok_hist, ConnectionError("boom")]):
                    report = build_case_period_report(input_path)

            self.assertEqual(len(report), 2)
            self.assertEqual(report.iloc[0]["end_date"], "2026-01-22")
            self.assertIn("ConnectionError", report.iloc[1]["error"])

    def test_apply_report_to_workbook_updates_end_date_and_max_profit(self):
        with TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "stock.xlsx"
            input_df = pd.DataFrame(
                [
                    {"case_id": "案例001", "标的名称": "五洲新春", "开始日期": "2025-11-05", "结束日期": "", "收益率(%)": ""},
                    {"case_id": "案例007", "标的名称": "通富微电", "开始日期": "2026-02-26", "结束日期": "", "收益率(%)": ""},
                    {"case_id": "案例005", "标的名称": "智普", "开始日期": "2026-02-07", "结束日期": "", "收益率(%)": ""},
                ]
            )
            with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
                input_df.to_excel(writer, sheet_name="案例库", index=False)

            report_df = pd.DataFrame(
                [
                    {"case_id": "案例001", "stock_name": "五洲新春", "end_date": "2026-01-22", "max_profit_pct": 130.48, "error": ""},
                    {"case_id": "案例007", "stock_name": "通富微电", "end_date": "2026-02-27", "max_profit_pct": 1.36, "error": ""},
                    {"case_id": "案例005", "stock_name": "智普", "end_date": "", "max_profit_pct": "", "error": "stock_code_not_found"},
                ]
            )

            updated_rows = apply_report_to_workbook(workbook_path, report_df, excluded_names={"通富微电"})

            self.assertEqual(updated_rows, 1)
            wb = load_workbook(workbook_path, data_only=True)
            ws = wb["案例库"]
            headers = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
            end_col = headers["结束日期"]
            profit_col = headers["收益率(%)"]
            self.assertEqual(str(ws.cell(2, end_col).value)[:10], "2026-01-22")
            self.assertAlmostEqual(float(ws.cell(2, profit_col).value), 130.48)
            self.assertIsNone(ws.cell(3, end_col).value)
            self.assertIsNone(ws.cell(3, profit_col).value)
            self.assertIsNone(ws.cell(4, end_col).value)
            self.assertIsNone(ws.cell(4, profit_col).value)


if __name__ == "__main__":
    unittest.main()
